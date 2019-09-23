from openpyxl import load_workbook
from tools.get_path import *
from tools.read_config import ReadConfig
from tools.get_data import GetData
from tools.do_re import DoRegx
cf=eval(ReadConfig().read_config(config_path,'MODE','mode'))
class DoExcle:
    def __init__(self,filename):
        self.filename=filename

    def get_headers(self,sheetname):
        wb = load_workbook(self.filename)
        sheet=wb[sheetname]
        headers=[]
        for j in range(1,sheet.max_column-1):
            headers.append(sheet.cell(1,j).value)
        return headers
    def get_data(self):
        wb = load_workbook(self.filename)
        data = []
        for key in cf:
            sheetname=key
            sheet=wb[sheetname]
            sheet_data=[]
            if cf[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    for j in range(1,sheet.max_column-2):
                        row_data[self.get_headers(sheetname)[j-1]]=sheet.cell(i,j).value
                    row_data['sheetname']=key
                    # 将已经产生确定值的变量批量替换，需要执行用例才产生确定值的变量不能替换
                    row_data = DoRegx.do_regx('\$\{(.*?)\}', str(row_data))
                    sheet_data.append(eval(row_data))
            else:
                for case_id in cf[key]:
                    row_data={}
                    for j in range(1,sheet.max_column-2):
                        row_data[self.get_headers(sheetname)[j-1]] = sheet.cell(case_id+1, j).value
                    row_data['sheetname'] = key
                    # 将已经产生确定值的变量批量替换，需要执行用例才产生确定值的变量不能替换
                    row_data=DoRegx.do_regx('\$\{(.*?)\}',str(row_data))
                    sheet_data.append(eval(row_data))
            data+=sheet_data
        return data

    def write_back(self,sheetname,i,j,value):
        wb = load_workbook(self.filename)
        sheet=wb[sheetname]
        sheet.cell(i,j).value=value
        wb.save(self.filename)
        wb.close()
if __name__ == '__main__':
    data=DoExcle(test_data_path).get_data()
    print(data)
